"""
Genera el seed JSON de la PWA (data/vocabulario.json, data/frases.json) a
partir del Excel original. A diferencia de Herramientas_HTML, aqui la fuente
de verdad NO es el .msapp -- las filas de Vocabulario/Frases viven en el
Excel (conectado como fuente de datos "Vocabulario_Ingles" en Power Apps),
el .msapp solo trae el esquema, no los datos.

Fuente: APP_PowerApps/Base de datos/Vocabulario_Ingles.xlsx
        Hoja "Vocabulario": Num_general, Lista, Palabra_Ing, Palabra_Esp,
                             Aprendida (Si/No), Palabra_en_contexto, __PowerAppsId__
        Hoja "Frases comunes": Num_general, Categoria, Frase_Ing, Frase_Esp,
                                Notas_de_uso, Aprendida (Si/No), __PowerAppsId__

Requiere: pip install openpyxl
Uso:      python tools/extract_data.py
Salida:   data/vocabulario.json, data/frases.json

Volver a ejecutar este script si el Excel original se actualiza y hay que
resincronizar el seed (OJO: esto NO toca lo que el usuario ya tenga guardado
en localStorage del navegador -- el seed solo se usa la primera vez que abre
la app en un navegador nuevo).
"""
import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / "APP_PowerApps" / "Base de datos" / "Vocabulario_Ingles.xlsx"
DATA_DIR = ROOT / "data"


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def sheet_rows(ws):
    headers = [clean(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        yield dict(zip(headers, row))


def write_json(name, data):
    DATA_DIR.mkdir(exist_ok=True)
    out_path = DATA_DIR / name
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  -> {out_path.relative_to(ROOT)}  ({len(data)} filas)")


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    print("Extrayendo vocabulario...")
    vocab_rows = []
    for row in sheet_rows(wb["Vocabulario"]):
        vocab_rows.append(
            {
                "id": clean(row.get("__PowerAppsId__")) or None,
                "lista": clean(row.get("Lista")),
                "palabra_ing": clean(row.get("Palabra_Ing")),
                "palabra_esp": clean(row.get("Palabra_Esp")),
                "aprendida": clean(row.get("Aprendida")).lower() == "si",
                "contexto": clean(row.get("Palabra_en_contexto")),
            }
        )
    vocab_rows = [r for r in vocab_rows if r["palabra_ing"] and r["palabra_esp"]]
    write_json("vocabulario.json", vocab_rows)

    print("Extrayendo frases comunes...")
    frase_rows = []
    for row in sheet_rows(wb["Frases comunes"]):
        frase_rows.append(
            {
                "id": clean(row.get("__PowerAppsId__")) or None,
                "categoria": clean(row.get("Categoria")),
                "frase_ing": clean(row.get("Frase_Ing")),
                "frase_esp": clean(row.get("Frase_Esp")),
                "notas_uso": clean(row.get("Notas_de_uso")),
                "aprendida": clean(row.get("Aprendida")).lower() == "si",
            }
        )
    # A diferencia de Vocabulario, aqui SI se conservan filas con Frase_Esp
    # vacio -- son frases en ingles que el usuario aun no ha traducido, y
    # descartarlas perderia ese trabajo en progreso (~62 de 123 filas).
    frase_rows = [r for r in frase_rows if r["frase_ing"]]
    write_json("frases.json", frase_rows)

    print("\nListo.")


if __name__ == "__main__":
    main()
